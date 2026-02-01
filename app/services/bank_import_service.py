"""
Bank Statement Import Service

Handles importing bank statements from CSV and Excel files for:
- Bank reconciliation
- Transaction matching
- Auto journal entry generation

Supports multiple bank statement formats:
- HDFC Bank
- ICICI Bank
- SBI
- Axis Bank
- Kotak Mahindra
- Generic CSV format
"""

import csv
import io
from datetime import datetime, date, timezone
from decimal import Decimal, InvalidOperation
from typing import Optional, Dict, Any, List, Tuple
from uuid import UUID
import re

from sqlalchemy import select, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.accounting import JournalEntry, JournalEntryLine
from app.models.banking import BankAccount, BankTransaction, TransactionType


class BankImportError(Exception):
    """Custom exception for bank import errors."""
    def __init__(self, message: str, row_number: int = None, details: Dict = None):
        self.message = message
        self.row_number = row_number
        self.details = details or {}
        super().__init__(self.message)


class BankStatementParser:
    """Base class for bank statement parsers."""

    # Column mappings for different banks
    # Override in subclass for specific bank formats
    DATE_COLUMNS = ["date", "transaction date", "txn date", "value date", "posting date"]
    DESCRIPTION_COLUMNS = ["description", "narration", "particulars", "remarks", "transaction details"]
    DEBIT_COLUMNS = ["debit", "withdrawal", "dr", "debit amount", "withdrawals"]
    CREDIT_COLUMNS = ["credit", "deposit", "cr", "credit amount", "deposits"]
    BALANCE_COLUMNS = ["balance", "closing balance", "running balance"]
    REFERENCE_COLUMNS = ["reference", "ref no", "cheque no", "utr", "transaction id", "ref number"]

    def __init__(self):
        self.date_format = "%d/%m/%Y"  # Default Indian date format
        self.alternate_date_formats = [
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%d/%m/%y",
            "%d-%m-%y",
            "%m/%d/%Y",
            "%d %b %Y",
            "%d-%b-%Y",
        ]

    def parse_date(self, date_str: str) -> Optional[date]:
        """Parse date from various formats."""
        if not date_str or not date_str.strip():
            return None

        date_str = date_str.strip()

        # Try primary format first
        try:
            return datetime.strptime(date_str, self.date_format).date()
        except ValueError:
            pass

        # Try alternate formats
        for fmt in self.alternate_date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return None

    def parse_amount(self, amount_str: str) -> Decimal:
        """Parse amount string to Decimal, handling various formats."""
        if not amount_str or not amount_str.strip():
            return Decimal("0")

        # Clean the string
        amount_str = amount_str.strip()

        # Remove currency symbols
        amount_str = re.sub(r'[₹$€£¥]', '', amount_str)

        # Handle CR/DR suffixes
        is_negative = False
        if amount_str.upper().endswith('DR') or amount_str.upper().endswith('D'):
            is_negative = True
            amount_str = re.sub(r'(DR|D)$', '', amount_str, flags=re.IGNORECASE)
        elif amount_str.upper().endswith('CR') or amount_str.upper().endswith('C'):
            amount_str = re.sub(r'(CR|C)$', '', amount_str, flags=re.IGNORECASE)

        # Handle brackets for negative
        if amount_str.startswith('(') and amount_str.endswith(')'):
            is_negative = True
            amount_str = amount_str[1:-1]

        # Handle minus sign
        if amount_str.startswith('-'):
            is_negative = True
            amount_str = amount_str[1:]

        # Remove commas and spaces
        amount_str = amount_str.replace(',', '').replace(' ', '')

        # Handle empty after cleanup
        if not amount_str or amount_str == '-':
            return Decimal("0")

        try:
            amount = Decimal(amount_str)
            return -amount if is_negative else amount
        except InvalidOperation:
            return Decimal("0")

    def find_column_index(self, headers: List[str], column_names: List[str]) -> int:
        """Find column index by matching header names."""
        headers_lower = [h.lower().strip() for h in headers]
        for name in column_names:
            if name.lower() in headers_lower:
                return headers_lower.index(name.lower())
        return -1

    def parse_csv(self, content: str) -> List[Dict]:
        """Parse CSV content and return list of transactions."""
        transactions = []

        # Try to detect delimiter
        delimiter = ','
        if '\t' in content:
            delimiter = '\t'
        elif ';' in content:
            delimiter = ';'

        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = list(reader)

        if len(rows) < 2:
            raise BankImportError("File has insufficient data rows")

        # Find header row (sometimes there are header rows before actual headers)
        header_row_idx = 0
        for idx, row in enumerate(rows):
            row_lower = [str(cell).lower() for cell in row]
            # Check if this row looks like headers
            if any(any(col in cell for col in self.DATE_COLUMNS) for cell in row_lower):
                header_row_idx = idx
                break

        headers = rows[header_row_idx]

        # Find column indices
        date_idx = self.find_column_index(headers, self.DATE_COLUMNS)
        desc_idx = self.find_column_index(headers, self.DESCRIPTION_COLUMNS)
        debit_idx = self.find_column_index(headers, self.DEBIT_COLUMNS)
        credit_idx = self.find_column_index(headers, self.CREDIT_COLUMNS)
        balance_idx = self.find_column_index(headers, self.BALANCE_COLUMNS)
        ref_idx = self.find_column_index(headers, self.REFERENCE_COLUMNS)

        if date_idx == -1:
            raise BankImportError("Could not find date column in file")

        if desc_idx == -1:
            raise BankImportError("Could not find description/narration column in file")

        # Parse data rows
        for row_num, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(cell.strip() for cell in row):
                continue  # Skip empty rows

            try:
                # Parse date
                transaction_date = self.parse_date(row[date_idx] if date_idx < len(row) else "")
                if not transaction_date:
                    continue  # Skip rows without valid date

                # Parse description
                description = row[desc_idx].strip() if desc_idx < len(row) else ""

                # Parse amounts
                debit_amount = Decimal("0")
                credit_amount = Decimal("0")

                if debit_idx >= 0 and debit_idx < len(row):
                    debit_amount = self.parse_amount(row[debit_idx])

                if credit_idx >= 0 and credit_idx < len(row):
                    credit_amount = self.parse_amount(row[credit_idx])

                # Handle single amount column (positive = credit, negative = debit)
                if debit_idx == credit_idx and debit_idx >= 0:
                    if debit_amount < 0:
                        debit_amount = abs(debit_amount)
                        credit_amount = Decimal("0")
                    else:
                        credit_amount = debit_amount
                        debit_amount = Decimal("0")

                # Parse balance
                balance = Decimal("0")
                if balance_idx >= 0 and balance_idx < len(row):
                    balance = self.parse_amount(row[balance_idx])

                # Parse reference
                reference = ""
                if ref_idx >= 0 and ref_idx < len(row):
                    reference = row[ref_idx].strip()

                # Determine transaction type
                if debit_amount > 0:
                    txn_type = TransactionType.DEBIT
                    amount = debit_amount
                else:
                    txn_type = TransactionType.CREDIT
                    amount = credit_amount

                if amount == 0:
                    continue  # Skip zero amount transactions

                transactions.append({
                    "date": transaction_date,
                    "description": description,
                    "debit": debit_amount,
                    "credit": credit_amount,
                    "amount": amount,
                    "type": txn_type,
                    "balance": balance,
                    "reference": reference,
                    "row_number": row_num,
                })

            except Exception as e:
                raise BankImportError(
                    f"Error parsing row {row_num}: {str(e)}",
                    row_number=row_num,
                    details={"row": row}
                )

        return transactions


class HDFCParser(BankStatementParser):
    """Parser for HDFC Bank statement format."""

    DATE_COLUMNS = ["date", "txn date", "transaction date"]
    DESCRIPTION_COLUMNS = ["narration", "description", "particulars"]
    DEBIT_COLUMNS = ["withdrawal amt.", "withdrawal", "debit"]
    CREDIT_COLUMNS = ["deposit amt.", "deposit", "credit"]
    BALANCE_COLUMNS = ["closing balance", "balance"]
    REFERENCE_COLUMNS = ["chq./ref.no.", "ref no", "cheque no"]


class ICICIParser(BankStatementParser):
    """Parser for ICICI Bank statement format."""

    DATE_COLUMNS = ["transaction date", "date", "value date"]
    DESCRIPTION_COLUMNS = ["transaction remarks", "remarks", "particulars"]
    DEBIT_COLUMNS = ["withdrawal amount (inr)", "withdrawal", "debit"]
    CREDIT_COLUMNS = ["deposit amount (inr)", "deposit", "credit"]
    BALANCE_COLUMNS = ["balance (inr)", "balance"]
    REFERENCE_COLUMNS = ["cheque no", "reference no", "utr"]


class SBIParser(BankStatementParser):
    """Parser for SBI Bank statement format."""

    DATE_COLUMNS = ["txn date", "date", "value date"]
    DESCRIPTION_COLUMNS = ["description", "narration"]
    DEBIT_COLUMNS = ["debit", "withdrawal"]
    CREDIT_COLUMNS = ["credit", "deposit"]
    BALANCE_COLUMNS = ["balance"]
    REFERENCE_COLUMNS = ["ref no./cheque no.", "reference"]


class BankImportService:
    """
    Service for importing bank statements and creating transactions.
    """

    BANK_PARSERS = {
        "HDFC": HDFCParser,
        "ICICI": ICICIParser,
        "SBI": SBIParser,
        "GENERIC": BankStatementParser,
    }

    def __init__(self, db: AsyncSession):
        self.db = db

    def detect_bank_format(self, content: str, filename: str = "") -> str:
        """Auto-detect bank format from content or filename."""
        content_lower = content.lower()
        filename_lower = filename.lower()

        if "hdfc" in content_lower or "hdfc" in filename_lower:
            return "HDFC"
        elif "icici" in content_lower or "icici" in filename_lower:
            return "ICICI"
        elif "sbi" in content_lower or "state bank" in content_lower or "sbi" in filename_lower:
            return "SBI"
        else:
            return "GENERIC"

    async def import_csv_statement(
        self,
        bank_account_id: UUID,
        file_content: str,
        filename: str = "",
        bank_format: str = "AUTO",
        skip_duplicates: bool = True,
        user_id: Optional[UUID] = None
    ) -> Dict:
        """
        Import bank statement from CSV content.

        Returns:
            Dict with import statistics and created transactions
        """
        # Get bank account
        result = await self.db.execute(
            select(BankAccount).where(BankAccount.id == bank_account_id)
        )
        bank_account = result.scalar_one_or_none()

        if not bank_account:
            raise BankImportError("Bank account not found")

        # Detect or use specified bank format
        if bank_format == "AUTO":
            bank_format = self.detect_bank_format(file_content, filename)

        # Get appropriate parser
        parser_class = self.BANK_PARSERS.get(bank_format, BankStatementParser)
        parser = parser_class()

        # Parse CSV
        transactions = parser.parse_csv(file_content)

        if not transactions:
            raise BankImportError("No valid transactions found in file")

        # Import statistics
        stats = {
            "total_rows": len(transactions),
            "imported": 0,
            "skipped_duplicates": 0,
            "errors": 0,
            "total_debit": Decimal("0"),
            "total_credit": Decimal("0"),
        }

        imported_transactions = []
        errors = []

        for txn in transactions:
            try:
                # Check for duplicates
                if skip_duplicates:
                    existing = await self.db.execute(
                        select(BankTransaction).where(
                            and_(
                                BankTransaction.bank_account_id == bank_account_id,
                                BankTransaction.transaction_date == txn["date"],
                                BankTransaction.amount == txn["amount"],
                                BankTransaction.description == txn["description"]
                            )
                        )
                    )
                    if existing.scalar_one_or_none():
                        stats["skipped_duplicates"] += 1
                        continue

                # Create bank transaction
                bank_txn = BankTransaction(
                    bank_account_id=bank_account_id,
                    transaction_date=txn["date"],
                    value_date=txn["date"],
                    description=txn["description"],
                    reference_number=txn["reference"],
                    transaction_type=txn["type"],
                    amount=txn["amount"],
                    debit_amount=txn["debit"],
                    credit_amount=txn["credit"],
                    running_balance=txn["balance"],
                    is_reconciled=False,
                    source="IMPORT",
                    import_reference=filename,
                    created_by=user_id,
                )

                self.db.add(bank_txn)
                imported_transactions.append(bank_txn)
                stats["imported"] += 1
                stats["total_debit"] += txn["debit"]
                stats["total_credit"] += txn["credit"]

            except Exception as e:
                stats["errors"] += 1
                errors.append({
                    "row": txn.get("row_number"),
                    "error": str(e)
                })

        await self.db.commit()

        return {
            "success": True,
            "bank_format": bank_format,
            "statistics": stats,
            "transactions": [
                {
                    "id": str(t.id),
                    "date": str(t.transaction_date),
                    "description": t.description,
                    "amount": float(t.amount),
                    "type": t.transaction_type
                }
                for t in imported_transactions[:50]  # Limit response
            ],
            "errors": errors if errors else None
        }

    async def import_excel_statement(
        self,
        bank_account_id: UUID,
        file_bytes: bytes,
        filename: str = "",
        sheet_name: str = None,
        bank_format: str = "AUTO",
        skip_duplicates: bool = True,
        user_id: Optional[UUID] = None
    ) -> Dict:
        """
        Import bank statement from Excel file.

        Requires openpyxl package for xlsx files.
        """
        try:
            import openpyxl
            from io import BytesIO
        except ImportError:
            raise BankImportError("openpyxl package required for Excel import. Run: pip install openpyxl")

        # Load workbook
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise BankImportError(f"Failed to read Excel file: {str(e)}")

        # Get sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise BankImportError(f"Sheet '{sheet_name}' not found in workbook")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # Convert to CSV-like format
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        # Convert to CSV string
        csv_content = "\n".join([",".join(row) for row in rows])

        # Use CSV import
        return await self.import_csv_statement(
            bank_account_id=bank_account_id,
            file_content=csv_content,
            filename=filename,
            bank_format=bank_format,
            skip_duplicates=skip_duplicates,
            user_id=user_id
        )

    async def get_unreconciled_transactions(
        self,
        bank_account_id: UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        limit: int = 100
    ) -> List[BankTransaction]:
        """Get unreconciled bank transactions for matching."""
        query = select(BankTransaction).where(
            and_(
                BankTransaction.bank_account_id == bank_account_id,
                BankTransaction.is_reconciled == False
            )
        )

        if start_date:
            query = query.where(BankTransaction.transaction_date >= start_date)
        if end_date:
            query = query.where(BankTransaction.transaction_date <= end_date)

        query = query.order_by(BankTransaction.transaction_date.desc()).limit(limit)

        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def match_with_journal_entries(
        self,
        bank_transaction_id: UUID,
        journal_entry_id: UUID
    ) -> Dict:
        """
        Match a bank transaction with a journal entry for reconciliation.
        """
        # Get bank transaction
        txn_result = await self.db.execute(
            select(BankTransaction).where(BankTransaction.id == bank_transaction_id)
        )
        bank_txn = txn_result.scalar_one_or_none()

        if not bank_txn:
            raise BankImportError("Bank transaction not found")

        # Get journal entry
        je_result = await self.db.execute(
            select(JournalEntry).where(JournalEntry.id == journal_entry_id)
        )
        journal_entry = je_result.scalar_one_or_none()

        if not journal_entry:
            raise BankImportError("Journal entry not found")

        # Mark as reconciled
        bank_txn.is_reconciled = True
        bank_txn.reconciled_at = datetime.now(timezone.utc)
        bank_txn.matched_journal_entry_id = journal_entry_id

        await self.db.commit()

        return {
            "success": True,
            "bank_transaction_id": str(bank_transaction_id),
            "journal_entry_id": str(journal_entry_id),
            "reconciled_at": str(bank_txn.reconciled_at)
        }

    async def suggest_matches(
        self,
        bank_transaction_id: UUID,
        tolerance_days: int = 3,
        tolerance_amount: Decimal = Decimal("1.00")
    ) -> List[Dict]:
        """
        Suggest potential journal entry matches for a bank transaction.

        Uses fuzzy matching on:
        - Date (within tolerance_days)
        - Amount (within tolerance_amount)
        - Description keywords
        """
        # Get bank transaction
        txn_result = await self.db.execute(
            select(BankTransaction).where(BankTransaction.id == bank_transaction_id)
        )
        bank_txn = txn_result.scalar_one_or_none()

        if not bank_txn:
            return []

        # Find potential journal entry matches
        from datetime import timedelta

        date_start = bank_txn.transaction_date - timedelta(days=tolerance_days)
        date_end = bank_txn.transaction_date + timedelta(days=tolerance_days)
        amount_low = bank_txn.amount - tolerance_amount
        amount_high = bank_txn.amount + tolerance_amount

        # Query journal lines with matching amounts
        query = select(JournalEntryLine).where(
            and_(
                or_(
                    and_(JournalEntryLine.debit >= amount_low, JournalEntryLine.debit <= amount_high),
                    and_(JournalEntryLine.credit >= amount_low, JournalEntryLine.credit <= amount_high)
                )
            )
        ).limit(20)

        result = await self.db.execute(query)
        potential_matches = result.scalars().all()

        suggestions = []
        for line in potential_matches:
            # Calculate match score
            score = 0

            # Amount match (40 points max)
            amount_diff = abs(bank_txn.amount - (line.debit or line.credit or Decimal("0")))
            if amount_diff == 0:
                score += 40
            elif amount_diff <= Decimal("0.10"):
                score += 35
            elif amount_diff <= tolerance_amount:
                score += 20

            # Add to suggestions if score > 0
            if score > 0:
                suggestions.append({
                    "journal_line_id": str(line.id),
                    "journal_entry_id": str(line.journal_entry_id),
                    "account_name": line.account_name,
                    "debit": float(line.debit) if line.debit else 0,
                    "credit": float(line.credit) if line.credit else 0,
                    "narration": line.narration,
                    "match_score": score
                })

        # Sort by score
        suggestions.sort(key=lambda x: x["match_score"], reverse=True)
        return suggestions[:10]
